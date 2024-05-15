import openpyxl.workbook
import pandas as pd
import os
import fnmatch
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import openpyxl

def match_with_wildcards(word, blacklist):
    for pattern in blacklist:
        if fnmatch.fnmatchcase(word, pattern):
            return True
    return False

def make_into_list(file):
    try:
        # Arbeitsmappe öffnen und Blatt auswählen
        wb = openpyxl.load_workbook(f'output/filtered_{file}')
        sheet = wb['Sheet']
        

        # Tabelle definieren
        tab = Table(displayName="Table1", ref=sheet.calculate_dimension())

        # Standardstil zur Tabelle hinzufügen
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style

        # Tabelle zum Blatt hinzufügen
        sheet.add_table(tab)

        # Arbeitsmappe speichern
        wb.save(f'output/filtered_{file}')
            
    except Exception as e:
        print(f"Fehler: Konnte keine Tabelle zur Datei {f'output/filtered_{file}'} hinzufügen. {str(e)}")

def auto_adjust_columns(workbook):
    for worksheet in workbook.worksheets:
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column  # Get the column name
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[get_column_letter(column)].width = adjusted_width


# Get list of all blacklist files in the blacklist directory
blacklist_files = os.listdir('blacklists')

# Create a dictionary where the keys are the names of the blacklist files (i.e., column names),
# and the values are the corresponding blacklisted words
blacklist = {}
for file in blacklist_files:
    if file.endswith('.txt'):
        with open(f'blacklists/{file}', 'r', encoding='utf-8') as f:
            blacklist[file[:-4]] = f.read().splitlines()


# Get list of all Excel files in the input directory
input_files = os.listdir('input')

for file in input_files:
    if file.endswith('.xlsx'):
        # Load the Excel file
        df = pd.read_excel(f'input/{file}')

        # Apply the filter to specified column for each blacklisted word
        for column in df.columns:
            df = df[~df[column].astype(str).str.startswith('=')]

        for column, words in blacklist.items():
            if column in df.columns:
                print(f'Column {column} found in file {file}. Filtering column...')
                df[column] = df[column].fillna('')  # replace NaNs with empty string
                df = df[~df[column].apply(lambda x: match_with_wildcards(str(x), words))]
            else:
                print(f'Column {column} not found in file {file}. Ignoring this column.')

        # Write the filtered data to a new Excel file in the output directory
        wb = openpyxl.Workbook()
        ws = wb.active
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        auto_adjust_columns(wb)
        wb.save(f'output/filtered_{file}')
        make_into_list(file)
