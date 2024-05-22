import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk
import openpyxl
import pandas as pd
import os
import fnmatch
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import subprocess
import threading
import configparser
import argparse

def match_with_wildcards(word, patterns):
    for pattern in patterns:
        if fnmatch.fnmatchcase(word, pattern):
            return True
    return False

def make_into_list(wb):
    try:
        log("Converting to list...")
        sheet = wb['Sheet']
        tab = Table(displayName="Table1", ref=sheet.calculate_dimension())
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        sheet.add_table(tab)
    except Exception as e:
        log(f"Error: Could not add table to file. {str(e)}")

def auto_adjust_columns(workbook):
    log("Auto adjusting column sizes...")
    for worksheet in workbook.worksheets:
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[get_column_letter(column)].width = adjusted_width

def process_files(input_dir, output_dir, blacklist_dir, whitelist_dir, blacklist_enabled, whitelist_enabled):
    try:
        # Read blacklist files
        blacklist_files = os.listdir(blacklist_dir)
        blacklist = {}
        for file in blacklist_files:
            if file.endswith('.txt'):
                with open(os.path.join(blacklist_dir, file), 'r', encoding='utf-8') as f:
                    contents = f.read().splitlines()
                    if contents:  # Ignore empty files
                        blacklist[file[:-4]] = contents

        # Read whitelist files
        whitelist_files = os.listdir(whitelist_dir)
        whitelist = {}
        for file in whitelist_files:
            if file.endswith('.txt'):
                with open(os.path.join(whitelist_dir, file), 'r', encoding='utf-8') as f:
                    contents = f.read().splitlines()
                    if contents:  # Ignore empty files
                        whitelist[file[:-4]] = contents

        # Process input Excel files
        input_files = os.listdir(input_dir)
        total_files = len([file for file in input_files if file.endswith('.xlsx')])
        processed_files = 0

        for file in input_files:
            if file.endswith('.xlsx'):
                log(f'Found file {file} | Reading now...')
                df = pd.read_excel(os.path.join(input_dir, file))

                # Filter out rows where any column starts with '='
                for column in df.columns:
                    df = df[~df[column].astype(str).str.startswith('=')]

                if blacklist_enabled:
                    # Apply blacklist filters
                    for column, words in blacklist.items():
                        if column in df.columns:
                            log(f'Column {column} found in file {file}. Filtering column with blacklist...')
                            df[column] = df[column].fillna('')  # replace NaNs with empty string
                            initial_row_count = df.shape[0]
                            df = df[~df[column].apply(lambda x: match_with_wildcards(str(x), words))]
                            filtered_row_count = df.shape[0]
                            log(f'Blacklist filtering on column {column} reduced rows from {initial_row_count} to {filtered_row_count}')
                        else:
                            log(f'Column {column} not found in file {file}. Ignoring this column.')

                if whitelist_enabled:
                    # Apply whitelist filters
                    for column, words in whitelist.items():
                        if column in df.columns:
                            log(f'Column {column} found in file {file}. Applying whitelist...')
                            initial_row_count = df.shape[0]
                            df = df[df[column].apply(lambda x: match_with_wildcards(str(x), words))]
                            filtered_row_count = df.shape[0]
                            log(f'Whitelist filtering on column {column} reduced rows from {initial_row_count} to {filtered_row_count}')
                        else:
                            log(f'Column {column} not found in file {file}. Ignoring this column.')

                # Check if dataframe is empty after filtering
                if df.empty:
                    log(f'No rows remaining after filtering for file {file}.')
                else:
                    # Write the filtered data to a new Excel file in the output directory
                    log("Writing filtered data to workbook.")
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    for r in dataframe_to_rows(df, index=False, header=True):
                        ws.append(r)
                    auto_adjust_columns(wb)
                    make_into_list(wb)
                    wb.save(os.path.join(output_dir, f'filtered_{file}'))
                
                processed_files += 1
                update_progress(processed_files, total_files)

        log("Processing complete.")
    except Exception as e:
        log(f"An error occurred: {str(e)}")
        if mode == "gui":
            messagebox.showerror("Error", f"An error occurred: {str(e)}")

def create_empty_lists(input_dir, blacklist_dir, whitelist_dir):
    try:
        input_files = [file for file in os.listdir(input_dir) if file.endswith('.xlsx')]
        if not input_files:
            log("No Excel files found in input directory.")
            return
        
        # Read columns from the first Excel file
        df = pd.read_excel(os.path.join(input_dir, input_files[0]))
        columns = df.columns.tolist()
        
        # Create empty blacklist and whitelist files for each column if they don't already exist
        for column in columns:
            blacklist_file = os.path.join(blacklist_dir, f'{column}.txt')
            whitelist_file = os.path.join(whitelist_dir, f'{column}.txt')
            if not os.path.exists(blacklist_file):
                with open(blacklist_file, 'w', encoding='utf-8') as f:
                    pass
            else:
                log(f"Blacklist file for column '{column}' already exists. Skipping creation.")
            if not os.path.exists(whitelist_file):
                with open(whitelist_file, 'w', encoding='utf-8') as f:
                    pass
            else:
                log(f"Whitelist file for column '{column}' already exists. Skipping creation.")
        
        log("Empty blacklist and whitelist files created where they did not exist.")
    except Exception as e:
        log(f"An error occurred while creating empty lists: {str(e)}")
        if mode == "gui":
            messagebox.showerror("Error", f"An error occurred while creating empty lists: {str(e)}")

def log(message):
    if mode == "gui":
        log_box.config(state=tk.NORMAL)
        log_box.insert(tk.END, message + "\n")
        log_box.config(state=tk.DISABLED)
        log_box.see(tk.END)
        root.update_idletasks()  # Ensure the GUI is updated
    else:
        print(message)

def update_progress(processed, total):
    if mode == "gui":
        progress_var.set((processed / total) * 100)
        root.update_idletasks()  # Ensure the progress bar is updated

def open_directory(path):
    if os.path.isdir(path):
        if os.name == 'nt':
            os.startfile(path)
        elif os.name == 'posix':
            subprocess.call(['open', path])

def start_processing():
    save_settings()
    threading.Thread(target=process_files, args=(input_dir, output_dir, blacklist_dir, whitelist_dir, blacklist_var.get(), whitelist_var.get())).start()

def create_empty_lists_from_columns():
    save_settings()
    threading.Thread(target=create_empty_lists, args=(input_dir, blacklist_dir, whitelist_dir)).start()

def load_settings():
    config = configparser.ConfigParser()
    if os.path.exists('settings.ini'):
        config.read('settings.ini')
        blacklist_var.set(config.getboolean('Filters', 'blacklist', fallback=True))
        whitelist_var.set(config.getboolean('Filters', 'whitelist', fallback=True))

def save_settings():
    config = configparser.ConfigParser()
    config['Filters'] = {
        'blacklist': str(blacklist_var.get()),
        'whitelist': str(whitelist_var.get())
    }
    with open('settings.ini', 'w') as configfile:
        config.write(configfile)

def main(args):
    global mode, blacklist_var, whitelist_var, input_dir, output_dir, blacklist_dir, whitelist_dir
    mode = args.mode
    blacklist_var = args.blacklist
    whitelist_var = args.whitelist
    input_dir = args.input_dir
    output_dir = args.output_dir
    blacklist_dir = args.blacklist_dir
    whitelist_dir = args.whitelist_dir

    if mode == "gui":
        # Create the main application window
        global root, log_box, progress_var
        root = tk.Tk()
        root.title("Excel File Processor")
        root.configure(bg='#2F3136')  # Discord dark mode background

        # Set the style for the ttk widgets
        style = ttk.Style(root)
        style.theme_use('clam')

        # Define colors
        dark_gray = '#2F3136'
        medium_gray = '#36393F'
        light_gray = '#B9BBBE'
        highlight_blue = '#7289DA'
        white = '#FFFFFF'

        # Configure styles
        style.configure('TLabel', font=('Calibri', 16), background=dark_gray, foreground=white)
        style.configure('TButton', font=('Calibri', 16), background='gray', foreground=white, borderwidth=0)
        style.map('TButton', background=[('active', highlight_blue)])
        style.configure('TCheckbutton', font=('Calibri', 12), background=dark_gray, foreground=white)
        style.map('TCheckbutton', background=[('active', medium_gray)])
        style.configure('TProgressbar', thickness=20)
        style.configure('custom.Horizontal.TProgressbar', troughcolor=medium_gray, background=highlight_blue)

        # Create and place the input directory widgets
        ttk.Label(root, text="Input Directory").grid(row=0, column=0, padx=10, pady=5, sticky="w")
        ttk.Button(root, text="Open", command=lambda: open_directory(input_dir)).grid(row=0, column=2, padx=10, pady=5)

        # Create and place the output directory widgets
        ttk.Label(root, text="Output Directory").grid(row=1, column=0, padx=10, pady=5, sticky="w")
        ttk.Button(root, text="Open", command=lambda: open_directory(output_dir)).grid(row=1, column=2, padx=10, pady=5)

        # Create and place the blacklist directory widgets
        ttk.Label(root, text="Blacklist Directory").grid(row=2, column=0, padx=10, pady=5, sticky="w")
        ttk.Button(root, text="Open", command=lambda: open_directory(blacklist_dir)).grid(row=2, column=2, padx=10, pady=5)

        # Create and place the whitelist directory widgets
        ttk.Label(root, text="Whitelist Directory").grid(row=3, column=0, padx=10, pady=5, sticky="w")
        ttk.Button(root, text="Open", command=lambda: open_directory(whitelist_dir)).grid(row=3, column=2, padx=10, pady=5)

        # Create and place the filter checkboxes
        blacklist_var = tk.BooleanVar(value=blacklist_var)
        whitelist_var = tk.BooleanVar(value=whitelist_var)
        ttk.Checkbutton(root, text="Enable Blacklist Filter", variable=blacklist_var).grid(row=4, column=0, columnspan=3, pady=5, sticky="w")
        ttk.Checkbutton(root, text="Enable Whitelist Filter", variable=whitelist_var).grid(row=5, column=0, columnspan=3, pady=5, sticky="w")

        # Create and place the run button
        ttk.Button(root, text="Process Files", command=start_processing).grid(row=6, column=0, columnspan=3, pady=10)

        # Create and place the create empty lists button
        ttk.Button(root, text="Create Empty Lists", command=create_empty_lists_from_columns).grid(row=4, column=2, columnspan=1, pady=10)

        # Create and place the log box
        log_box = scrolledtext.ScrolledText(root, width=60, height=10, state='disabled', font=('Arial', 10), bg=dark_gray, fg=white, insertbackground=white)
        log_box.grid(row=8, column=0, columnspan=3, padx=10, pady=10)

        # Create and place the progress bar
        progress_var = tk.DoubleVar()
        progress_bar = ttk.Progressbar(root, variable=progress_var, maximum=100, style='custom.Horizontal.TProgressbar')
        progress_bar.grid(row=9, column=0, columnspan=3, padx=10, pady=10)

        # Add padding to all widgets for a cleaner look
        for widget in root.winfo_children():
            widget.grid_configure(padx=10, pady=10)

        # Load settings on startup
        load_settings()

        # Start the Tkinter event loop
        root.mainloop()
    else:
        process_files(input_dir, output_dir, blacklist_dir, whitelist_dir, blacklist_var, whitelist_var)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Process Excel files with blacklist and whitelist filtering.")
    parser.add_argument("--mode", type=str, choices=["gui", "console"], default="gui", help="Mode to run the application in (gui or console).")
    parser.add_argument("--blacklist", action='store_true', help="Enable blacklist filtering.")
    parser.add_argument("--no-blacklist", dest='blacklist', action='store_false', help="Disable blacklist filtering.")
    parser.add_argument("--whitelist", action='store_true', help="Enable whitelist filtering.")
    parser.add_argument("--no-whitelist", dest='whitelist', action='store_false', help="Disable whitelist filtering.")
    parser.add_argument("--input_dir", type=str, default="input", help="Input directory.")
    parser.add_argument("--output_dir", type=str, default="output", help="Output directory.")
    parser.add_argument("--blacklist_dir", type=str, default="blacklists", help="Blacklist directory.")
    parser.add_argument("--whitelist_dir", type=str, default="whitelists", help="Whitelist directory.")
    parser.set_defaults(blacklist=True, whitelist=True)
    args = parser.parse_args()
    main(args)
